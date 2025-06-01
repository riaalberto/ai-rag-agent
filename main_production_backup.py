# main_production.py - Versión optimizada para Railway con Supabase + EXCEL SUPPORT
import os
import ssl
import socket
import urllib3
from urllib3.util.ssl_ import create_urllib3_context
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import uvicorn
from datetime import datetime
import json
import re
import uuid
import PyPDF2
import io

# 🆕 NUEVO: Importación para Excel
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
    print("✅ Excel processor available (openpyxl)")
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Excel processor not available - install openpyxl")

# Configuración SSL para Pinecone
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
ctx = create_urllib3_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE
ctx.set_ciphers('ECDHE+AESGCM:ECDHE+CHACHA20:DHE+AESGCM:DHE+CHACHA20:!aNULL:!MD5:!DSS')
ssl._create_default_https_context = lambda: ctx
socket.setdefaulttimeout(30)

# Variables de entorno
PINECONE_API_KEY = os.getenv("PINECONE_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

print("🔍 DEBUG: Starting main_production.py with Supabase integration")

# IMPORTAR CONFIGURACIÓN DE DATABASE.PY (CON SUPABASE HARDCODEADO)
try:
    from database import (
        get_db_cursor, 
        get_db_connection, 
        get_documents, 
        get_document_content,
        create_conversation,
        DATABASE_CONFIG
    )
    print(f"✅ SUCCESS: Imported database functions")
    print(f"🔍 DEBUG: Using database host: {DATABASE_CONFIG['host']}:{DATABASE_CONFIG['port']}")
    DATABASE_AVAILABLE = True
except ImportError as e:
    print(f"❌ ERROR: Failed to import database: {e}")
    DATABASE_AVAILABLE = False

# Configurar Pinecone
USE_PINECONE = PINECONE_API_KEY is not None
if USE_PINECONE:
    try:
        from pinecone import Pinecone
        pc = Pinecone(api_key=PINECONE_API_KEY)
        INDEX_NAME = "rag-cff-2048"
        pinecone_index = pc.Index(INDEX_NAME)
        print("✅ Pinecone configurado para producción")
    except Exception as e:
        print(f"⚠️ Error Pinecone: {e}")
        USE_PINECONE = False
        pinecone_index = None
else:
    pinecone_index = None

# Configurar OpenAI
USE_OPENAI = OPENAI_API_KEY is not None and OPENAI_API_KEY.startswith("sk-")
if USE_OPENAI:
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY)
    print("✅ OpenAI configurado para producción")

# Crear app FastAPI
app = FastAPI(
    title="🤖 RAG Service Production",
    description="Servicio RAG híbrido en producción con Railway + Supabase + Excel",
    version="6.1.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# CORS para permitir conexiones desde frontend - VERSIÓN CORREGIDA
cors_origins = os.getenv("CORS_ORIGINS", "*")
print(f"🔍 DEBUG: CORS_ORIGINS value: {cors_origins}")

# Configurar orígenes permitidos
if cors_origins == "*":
    allowed_origins = ["*"]
    print("🌐 CORS: Allowing all origins (*)")
else:
    allowed_origins = cors_origins.split(",")
    print(f"🌐 CORS: Allowing specific origins: {allowed_origins}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# Modelos
class ChatRequest(BaseModel):
    question: str
    user_id: str = "119f7084-be9e-416f-81d6-3ffeadb062d5"  # ✅ UUID VÁLIDO

class ChatResponse(BaseModel):
    id: str
    question: str
    answer: str
    sources: List[dict]
    timestamp: str
    user_id: str
    ai_model: str

class UploadResponse(BaseModel):
    success: bool
    document_id: str
    filename: str
    message: str
    file_size: int
    file_type: str

# 🆕 FUNCIONES PARA PROCESAMIENTO DE ARCHIVOS CON EXCEL

def extract_text_from_pdf(content: bytes) -> str:
    """Extraer texto de PDF"""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        print(f"❌ Error extracting PDF text: {e}")
        return ""

def extract_text_from_excel(content: bytes) -> str:
    """🆕 NUEVO: Extraer texto de archivos Excel (.xlsx, .xls)"""
    if not EXCEL_AVAILABLE:
        return "Error: openpyxl no está instalado para procesar archivos Excel"
    
    try:
        # Cargar el archivo Excel
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        
        print(f"📊 Procesando Excel con {len(workbook.sheetnames)} hojas")
        
        # Texto extraído
        extracted_text = "=== ARCHIVO EXCEL ===\n"
        extracted_text += f"Hojas disponibles: {', '.join(workbook.sheetnames)}\n\n"
        
        # Procesar cada hoja
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            extracted_text += f"=== HOJA: {sheet_name} ===\n"
            
            # Obtener dimensiones de datos
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            print(f"📋 Hoja '{sheet_name}': {max_row} filas x {max_col} columnas")
            
            # Extraer encabezados (primera fila)
            headers = []
            for col in range(1, min(max_col + 1, 50)):  # Limitar a 50 columnas
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value))
                else:
                    headers.append(f"Col{col}")
            
            if headers:
                extracted_text += f"COLUMNAS: {' | '.join(headers)}\n"
            
            # Extraer datos (limitar filas para evitar textos enormes)
            rows_to_process = min(max_row, 200)  # Máximo 200 filas
            data_rows = 0
            
            for row in range(2, rows_to_process + 1):  # Empezar desde fila 2 (después de headers)
                row_data = []
                has_data = False
                
                for col in range(1, min(max_col + 1, 50)):  # Limitar columnas
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value))
                        has_data = True
                    else:
                        row_data.append("")
                
                # Solo agregar filas que tengan datos
                if has_data:
                    extracted_text += f"FILA {row}: {' | '.join(row_data)}\n"
                    data_rows += 1
            
            # Información adicional si hay más filas
            if max_row > rows_to_process:
                extracted_text += f"\n... y {max_row - rows_to_process} filas adicionales\n"
            
            extracted_text += f"\nRESUMEN HOJA '{sheet_name}': {data_rows} filas con datos\n\n"
        
        # Metadatos finales
        extracted_text += "=== INFORMACIÓN DEL ARCHIVO ===\n"
        extracted_text += f"Total de hojas: {len(workbook.sheetnames)}\n"
        extracted_text += f"Hojas procesadas: {', '.join(workbook.sheetnames)}\n"
        
        print(f"✅ Excel procesado: {len(extracted_text)} caracteres extraídos")
        return extracted_text
        
    except Exception as e:
        error_msg = f"Error procesando archivo Excel: {str(e)}"
        print(f"❌ {error_msg}")
        return error_msg

def extract_text_from_file(content: bytes, content_type: str, filename: str) -> str:
    """🆕 MEJORADO: Extraer texto según tipo de archivo - CON SOPORTE EXCEL"""
    try:
        print(f"📄 Procesando: {filename} ({content_type})")
        
        # 🆕 EXCEL FILES - NUEVA FUNCIONALIDAD
        if (content_type in [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel'  # .xls
        ] or filename.lower().endswith(('.xlsx', '.xls'))):
            print("📊 Detectado archivo Excel")
            return extract_text_from_excel(content)
        
        # PDF Files
        elif content_type == "application/pdf" or filename.lower().endswith('.pdf'):
            print("📄 Detectado archivo PDF")
            return extract_text_from_pdf(content)
        
        # Text Files
        elif content_type == "text/plain" or filename.lower().endswith('.txt'):
            print("📝 Detectado archivo de texto")
            return content.decode('utf-8', errors='ignore')
        
        # Word Files (básico)
        elif (content_type in [
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'application/msword'
        ] or filename.lower().endswith(('.docx', '.doc'))):
            print("📝 Detectado archivo Word")
            # Por ahora, tratarlos como texto plano básico
            try:
                return content.decode('utf-8', errors='ignore')
            except:
                return f"Archivo Word detectado: {filename} - contenido no procesable automáticamente"
        
        # JSON Files
        elif content_type == "application/json" or filename.lower().endswith('.json'):
            print("📋 Detectado archivo JSON")
            try:
                json_data = json.loads(content.decode('utf-8'))
                return f"=== ARCHIVO JSON ===\n{json.dumps(json_data, indent=2, ensure_ascii=False)}"
            except:
                return f"Archivo JSON detectado: {filename} - formato inválido"
        
        # CSV Files
        elif content_type == "text/csv" or filename.lower().endswith('.csv'):
            print("📊 Detectado archivo CSV")
            try:
                csv_text = content.decode('utf-8', errors='ignore')
                return f"=== ARCHIVO CSV ===\n{csv_text}"
            except:
                return f"Archivo CSV detectado: {filename} - contenido no procesable"
        
        # Otros tipos
        else:
            print(f"⚠️ Tipo de archivo no específicamente soportado: {content_type}")
            # Intentar como texto plano
            try:
                return content.decode('utf-8', errors='ignore')
            except:
                return f"Archivo {filename} - contenido no procesable automáticamente"
                
    except Exception as e:
        error_msg = f"Error procesando archivo {filename}: {str(e)}"
        print(f"❌ {error_msg}")
        return error_msg

def save_document_to_supabase(document_id: str, user_id: str, filename: str, content: str, file_size: int):
    """Guardar documento usando database.py - VERSIÓN CORREGIDA"""
    try:
        print(f"📤 DEBUG: Saving to database using database.py functions")
        print(f"📤 DEBUG: Filename: {filename}")
        print(f"📤 DEBUG: Content length: {len(content)} chars")
        
        # USAR LA FUNCIÓN DE database.py QUE YA FUNCIONA
        from database import create_document
        
        result = create_document(
            name=filename,
            content=content,
            size=file_size,
            user_id=user_id,
            metadata={'document_id': document_id}
        )
        
        print(f"✅ SUCCESS: Document saved via database.py: {result}")
        return result
        
    except Exception as e:
        print(f"❌ Error saving via database.py: {e}")
        raise

# Sinónimos y funciones de búsqueda
SYNONYMS = {
    'vacaciones': ['vacaciones', 'vacación', 'descanso', 'días libres', 'ausencia', 'permiso', 'tiempo libre'],
    'política': ['política', 'políticas', 'norma', 'normas', 'regla', 'reglas', 'procedimiento'],
    'solicitar': ['solicitar', 'pedir', 'requerir', 'tramitar', 'gestionar', 'obtener'],
    'trabajo': ['trabajo', 'laboral', 'empleo', 'empresa', 'oficina'],
    'horario': ['horario', 'horarios', 'tiempo', 'horas', 'jornada'],
    'manual': ['manual', 'guía', 'instructivo', 'documentación'],
    'sistema': ['sistema', 'plataforma', 'herramienta', 'aplicación']
}

def expand_query_terms(question: str) -> set:
    """Expandir términos de búsqueda"""
    question_lower = question.lower()
    expanded_terms = set()
    
    original_words = re.findall(r'\b\w+\b', question_lower)
    expanded_terms.update(original_words)
    
    for word in original_words:
        for key, synonyms in SYNONYMS.items():
            if word in synonyms:
                expanded_terms.update(synonyms)
    
    stop_words = {
        'el', 'la', 'de', 'que', 'y', 'a', 'en', 'un', 'es', 'se', 'no', 'te', 'lo', 'le', 'da', 'su', 
        'por', 'son', 'con', 'para', 'como', 'las', 'del', 'los', 'una', 'mas', 'pero', 'sus', 'muy',
        'qué', 'cómo', 'cuál', 'dónde', 'cuándo', 'quién'
    }
    
    return expanded_terms - stop_words

def calculate_relevance(content: str, doc_name: str, question_terms: set) -> float:
    """Calcular relevancia"""
    content_lower = content.lower()
    doc_name_lower = doc_name.lower()
    
    content_words = set(re.findall(r'\b\w+\b', content_lower))
    exact_matches = len(question_terms.intersection(content_words))
    content_relevance = exact_matches / max(len(question_terms), 1) if question_terms else 0
    
    doc_words = set(re.findall(r'\b\w+\b', doc_name_lower.replace('_', ' ').replace('.', ' ')))
    title_matches = len(question_terms.intersection(doc_words))
    title_relevance = (title_matches / max(len(question_terms), 1)) * 0.3 if question_terms else 0
    
    frequency_bonus = 0
    for term in question_terms:
        frequency_bonus += content_lower.count(term) * 0.05
    
    document_type_bonus = 0
    if 'política' in doc_name_lower:
        document_type_bonus += 0.2
    if 'manual' in doc_name_lower:
        document_type_bonus += 0.15
    
    total_relevance = content_relevance + title_relevance + min(frequency_bonus, 0.3) + document_type_bonus
    return min(total_relevance, 1.0)

def production_search(question: str, user_id: str = None) -> List[dict]:
    """Búsqueda optimizada para producción usando Supabase"""
    try:
        print(f"🔍 DEBUG: Searching for question: {question}")
        print(f"🔍 DEBUG: Using user_id: {user_id}")
        
        if not DATABASE_AVAILABLE:
            print("❌ ERROR: Database not available")
            return []
            
        question_terms = expand_query_terms(question)
        print(f"🔍 DEBUG: Expanded terms: {question_terms}")
        
        # USAR FUNCIÓN DE database.py (CON SUPABASE)
        docs_data = get_documents(user_id)
        print(f"🔍 DEBUG: Found {len(docs_data)} documents from Supabase")
        
        if not docs_data:
            print("⚠️ WARNING: No documents found in Supabase")
            return []
        
        processed_docs = [doc for doc in docs_data if doc.get('status') == 'processed']
        print(f"🔍 DEBUG: {len(processed_docs)} processed documents")
        
        relevant_docs = []
        for doc in processed_docs:
            try:
                # USAR FUNCIÓN DE database.py (CON SUPABASE)
                doc_content = get_document_content(doc['id'], user_id)
                if not doc_content:
                    continue
                    
                content = doc_content['content']
                relevance = calculate_relevance(content, doc['name'], question_terms)
                
                if relevance > 0.01:
                    # Extraer mejor fragmento
                    sentences = content.split('.')
                    best_excerpt = ""
                    best_score = 0
                    
                    for sentence in sentences[:20]:
                        sentence_lower = sentence.lower()
                        sentence_words = set(re.findall(r'\b\w+\b', sentence_lower))
                        matches = len(question_terms.intersection(sentence_words))
                        
                        if matches > best_score:
                            best_score = matches
                            best_excerpt = sentence.strip()
                    
                    if not best_excerpt and sentences:
                        best_excerpt = sentences[0].strip()
                    
                    relevant_docs.append({
                        'document_id': doc['id'],
                        'document_name': doc['name'],
                        'content': content,
                        'excerpt': best_excerpt[:300] + "..." if len(best_excerpt) > 300 else best_excerpt,
                        'relevance': relevance
                    })
                    
            except Exception as e:
                print(f"❌ Error procesando documento: {e}")
                continue
        
        relevant_docs.sort(key=lambda x: x['relevance'], reverse=True)
        print(f"✅ SUCCESS: Found {len(relevant_docs)} relevant documents")
        return relevant_docs[:3]
        
    except Exception as e:
        print(f"❌ Error en búsqueda: {e}")
        return []

def generate_production_answer(question: str, relevant_docs: List[dict]) -> str:
    """Generar respuesta para producción"""
    if not relevant_docs:
        return "Lo siento, no encontré información relevante en los documentos disponibles."
    
    if USE_OPENAI:
        try:
            context = "\n\n".join([
                f"DOCUMENTO: {doc['document_name']}\nCONTENIDO: {doc['excerpt']}"
                for doc in relevant_docs[:2]
            ])
            
            prompt = f"""Basándote en estos documentos empresariales, responde la pregunta de manera precisa y detallada:

{context}

PREGUNTA: {question}

Responde en español, cita las fuentes y sé específico."""

            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "Eres un asistente experto en documentos empresariales."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=400,
                temperature=0.2
            )
            
            return response.choices[0].message.content.strip()
            
        except Exception as e:
            print(f"❌ Error OpenAI: {e}")
    
    # Respuesta local
    best_doc = relevant_docs[0]
    return f"Según el documento '{best_doc['document_name']}': {best_doc['excerpt']}"

# Rutas de producción
@app.get("/")
async def root():
    return {
        "message": "🤖 RAG Service Production - Railway Deployment + Excel Support",
        "status": "active",
        "environment": "production",
        "database": f"Supabase ({DATABASE_CONFIG['host']})" if DATABASE_AVAILABLE else "Disconnected",
        "vector_search": "Pinecone" if USE_PINECONE else "Disabled",
        "ai": "OpenAI GPT-3.5" if USE_OPENAI else "Local",
        "excel_support": "Available" if EXCEL_AVAILABLE else "Not Available",
        "version": "6.2.0-railway-supabase-excel"
    }

@app.get("/health")
async def health():
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "services": {
            "database": "connected" if DATABASE_AVAILABLE else "disconnected",
            "pinecone": "enabled" if USE_PINECONE else "disabled",
            "openai": "enabled" if USE_OPENAI else "enabled",
            "excel": "enabled" if EXCEL_AVAILABLE else "disabled"
        }
    }

# 🆕 ENDPOINT DE UPLOAD MEJORADO CON EXCEL
@app.post("/upload")
async def upload_document(
    file: UploadFile = File(...),
    user_id: str = "119f7084-be9e-416f-81d6-3ffeadb062d5"
):
    """🆕 MEJORADO: Subir y procesar documento - CON SOPORTE PARA EXCEL"""
    try:
        print(f"📤 DEBUG: Uploading file: {file.filename}")
        print(f"📤 DEBUG: Content type: {file.content_type}")
        print(f"📤 DEBUG: User ID: {user_id}")
        
        # 🆕 TIPOS DE ARCHIVO PERMITIDOS AMPLIADOS CON EXCEL
        allowed_types = [
            # PDFs
            'application/pdf',
            
            # Texto
            'text/plain',
            
            # Microsoft Office
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',  # .docx
            'application/msword',  # .doc
            
            # 🆕 EXCEL - NUEVOS TIPOS
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
            'application/vnd.ms-excel',  # .xls
            
            # Datos
            'application/json',
            'text/csv'
        ]
        
        # Verificar extensión también (más confiable que content-type)
        allowed_extensions = ['.pdf', '.txt', '.docx', '.doc', '.xlsx', '.xls', '.json', '.csv']
        file_extension = '.' + file.filename.lower().split('.')[-1] if '.' in file.filename else ''
        
        # Validación más flexible
        type_valid = file.content_type in allowed_types
        extension_valid = file_extension in allowed_extensions
        
        if not (type_valid or extension_valid):
            raise HTTPException(
                status_code=400, 
                detail=f"Tipo de archivo no soportado: {file.content_type}. "
                       f"Formatos permitidos: PDF, TXT, DOCX, DOC, XLSX, XLS, JSON, CSV"
            )
        
        print(f"✅ Archivo validado: tipo={type_valid}, extensión={extension_valid}")
        
        # Leer contenido
        content = await file.read()
        file_size = len(content)
        
        if file_size > 100 * 1024 * 1024:  # 100MB
            raise HTTPException(status_code=400, detail="Archivo demasiado grande. Máximo 100MB.")
        
        # 🆕 EXTRAER TEXTO CON SOPORTE MEJORADO PARA EXCEL
        text_content = extract_text_from_file(content, file.content_type, file.filename)
        
        if not text_content.strip():
            raise HTTPException(status_code=400, detail="No se pudo extraer texto del archivo.")
        
        print(f"📊 Texto extraído: {len(text_content)} caracteres")
        
        # Generar ID único
        document_id = str(uuid.uuid4())
        
        # Guardar en Supabase usando database.py
        save_document_to_supabase(
            document_id=document_id,
            user_id=user_id,
            filename=file.filename,
            content=text_content,
            file_size=file_size
        )
        
        print(f"✅ SUCCESS: Document uploaded: {document_id}")
        
        return UploadResponse(
            success=True,
            document_id=document_id,
            filename=file.filename,
            message="Documento subido y procesado exitosamente",
            file_size=file_size,
            file_type=file.content_type
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR uploading document: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")

@app.post("/chat")
async def production_chat(request: ChatRequest):
    """Chat endpoint para producción"""
    try:
        print(f"🔍 DEBUG: Chat request: {request.question}")
        print(f"🔍 DEBUG: Chat user_id: {request.user_id}")
        
        relevant_docs = production_search(request.question, request.user_id)
        answer = generate_production_answer(request.question, relevant_docs)
        
        sources = []
        for doc in relevant_docs:
            sources.append({
                'document': doc['document_name'],
                'excerpt': doc['excerpt'],
                'relevance': round(doc['relevance'], 3)
            })
        
        # USAR FUNCIÓN DE database.py (CON SUPABASE)
        if DATABASE_AVAILABLE:
            conv_result = create_conversation(
                question=request.question,
                answer=answer,
                sources=json.dumps(sources),
                user_id=request.user_id,
                ai_model='production-rag-railway-supabase-excel'
            )
        else:
            conv_result = {"id": "error", "timestamp": datetime.now()}
        
        return ChatResponse(
            id=conv_result['id'],
            question=request.question,
            answer=answer,
            sources=sources,
            timestamp=conv_result['timestamp'].isoformat() if hasattr(conv_result['timestamp'], 'isoformat') else str(conv_result['timestamp']),
            user_id=request.user_id,
            ai_model='production-rag-railway-supabase-excel'
        )
        
    except Exception as e:
        print(f"❌ ERROR in chat: {e}")
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@app.get("/documents")
async def get_production_documents(user_id: str = None):
    """Obtener documentos en producción desde Supabase"""
    try:
        print(f"🔍 DEBUG: Getting documents for user: {user_id}")
        
        if not DATABASE_AVAILABLE:
            print("❌ ERROR: Database not available")
            return []
            
        # USAR FUNCIÓN DE database.py (CON SUPABASE)
        docs = get_documents(user_id)
        print(f"✅ SUCCESS: Retrieved {len(docs)} documents from Supabase")
        
        return [
            {
                "id": doc['id'],
                "name": doc['name'],
                "size": f"{doc['size'] / (1024*1024):.1f} MB" if isinstance(doc['size'], int) else str(doc['size']),
                "status": doc['status'],
                "upload_date": doc['upload_date'].isoformat() if hasattr(doc['upload_date'], 'isoformat') else str(doc['upload_date'])
            }
            for doc in docs
        ]
    except Exception as e:
        print(f"❌ ERROR getting documents: {e}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)